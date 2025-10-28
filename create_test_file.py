#!/usr/bin/env python3
"""
Create a test Excel file with Dutch text for translation testing
"""

import openpyxl
from openpyxl import Workbook
from datetime import datetime

# Create a new workbook
wb = Workbook()

# First sheet: Medical Messages
ws1 = wb.active
ws1.title = "Medical Messages"

# Add headers
ws1['A1'] = "ID"
ws1['B1'] = "Patient"
ws1['C1'] = "Message"
ws1['D1'] = "Category"
ws1['E1'] = "Date"

# Add sample data (Dutch medical messages)
data = [
    [1, "Patient A", "Goede morgen, ik heb vandaag last van buikpijn.", "Symptoom", "2024-01-15"],
    [2, "Patient B", "De medicijnen werken goed, ik voel me veel beter.", "Feedback", "2024-01-16"],
    [3, "Patient C", "Wanneer is mijn volgende afspraak met de specialist?", "Vraag", "2024-01-17"],
    [4, "Patient D", "Ik heb mijn medicatie vergeten in te nemen gisteren.", "Medicatie", "2024-01-18"],
    [5, "Patient E", "De pijn is erger geworden sinds gisteravond.", "Symptoom", "2024-01-19"],
    [6, "Patient F", "Dank u voor de snelle reactie op mijn eerdere vraag.", "Feedback", "2024-01-20"],
    [7, "Patient G", "Kan ik mijn dieet aanpassen volgens uw advies?", "Vraag", "2024-01-21"],
    [8, "Patient H", "Ik heb bijwerkingen van de nieuwe medicatie.", "Medicatie", "2024-01-22"],
    [9, "Patient I", "De behandeling heeft echt geholpen, bedankt!", "Feedback", "2024-01-23"],
    [10, "Patient J", "Moet ik de dosering verhogen zoals besproken?", "Vraag", "2024-01-24"],
]

for row_idx, row_data in enumerate(data, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        ws1.cell(row=row_idx, column=col_idx, value=value)

# Second sheet: Common Phrases
ws2 = wb.create_sheet("Common Phrases")
ws2['A1'] = "Dutch"
ws2['B1'] = "Context"

phrases = [
    ["Goedemorgen, hoe gaat het met u?", "Greeting"],
    ["Ik begrijp uw zorgen.", "Empathy"],
    ["Neem contact met ons op als u vragen heeft.", "Support"],
    ["Uw gezondheid is onze prioriteit.", "Care"],
    ["De testresultaten zijn binnen.", "Results"],
    ["Heeft u last van bijwerkingen?", "Question"],
    ["Neem de medicatie in volgens voorschrift.", "Instruction"],
    ["Maak een afspraak via onze website.", "Action"],
]

for row_idx, row_data in enumerate(phrases, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        ws2.cell(row=row_idx, column=col_idx, value=value)

# Third sheet: Product Descriptions
ws3 = wb.create_sheet("Product Info")
ws3['A1'] = "Product"
ws3['B1'] = "Description"
ws3['C1'] = "Price"

products = [
    ["Pijnstiller", "Effectieve pijnverlichting voor hoofdpijn en spierpijn.", "€8.99"],
    ["Vitamine D", "Ondersteunt sterke botten en een gezond immuunsysteem.", "€12.50"],
    ["Ontstekingsremmer", "Vermindert ontsteking en zwelling.", "€15.75"],
    ["Allergietabletten", "Verlicht symptomen van hooikoorts en allergieën.", "€9.99"],
    ["Slaaptabletten", "Helpt bij het vinden van een goede nachtrust.", "€11.25"],
]

for row_idx, row_data in enumerate(products, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        ws3.cell(row=row_idx, column=col_idx, value=value)

# Save the workbook
filename = "test_translation_file.xlsx"
wb.save(filename)

print(f"✅ Test file created: {filename}")
print(f"📊 Contains 3 sheets:")
print(f"   1. Medical Messages ({len(data)} rows)")
print(f"   2. Common Phrases ({len(phrases)} rows)")
print(f"   3. Product Info ({len(products)} rows)")
print(f"\n🌍 All text is in Dutch (Nederlands) - ready for translation!")
